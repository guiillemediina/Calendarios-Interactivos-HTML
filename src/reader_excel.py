from typing import List
from openpyxl import load_workbook
from .models import Event


def cargar_eventos_desde_excel(path: str) -> List[Event]:
    """
    Lee eventos desde un archivo Excel (.xlsx) y devuelve una lista de Event.

    Se espera una hoja con encabezados en la primera fila:
    titulo, descripcion, categoria, fecha_inicio, fecha_fin, hora_inicio, hora_fin
    """

    # 1) Intentar abrir el libro de Excel
    try:
        wb = load_workbook(path, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"No se ha encontrado el archivo Excel: {path}")
    except Exception as e:
        raise ValueError(f"Error al abrir el Excel: {e}")

    # 2) Usar la hoja activa (normalmente la primera)
    ws = wb.active

    # 3) Leer los encabezados de la primera fila y mapear nombre_columna -> índice_columna
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            headers[str(header_value)] = col

    # 4) Comprobar que están las columnas obligatorias
    columnas_obligatorias = ["titulo", "descripcion", "categoria", "fecha_inicio"]
    for col_name in columnas_obligatorias:
        if col_name not in headers:
            raise ValueError(f"Falta la columna obligatoria en Excel: {col_name}")

    eventos: List[Event] = []
    errors: List[str] = []

    # 5) Recorrer filas de datos (desde la 2 en adelante)
    for row in range(2, ws.max_row + 1):
        # Si la fila está vacía, la saltamos
        if _fila_vacia(ws, row):
            continue

        data = {}

        # Para cada campo que nos interesa, leemos el valor de la celda
        for field in ["titulo", "descripcion", "categoria",
                      "fecha_inicio", "fecha_fin", "hora_inicio", "hora_fin"]:
            col = headers.get(field)
            if col is not None:
                cell_value = ws.cell(row=row, column=col).value
                data[field] = cell_value

        # Intentar crear el Event desde el dict
        try:
            evento = Event.desde_dict(data)
            eventos.append(evento)
        except Exception as e:
            errors.append(f"Fila {row}: {e}")

    # 6) Si hubo errores, los mostramos todos juntos
    if errors:
        mensajes = "\n".join(errors)
        raise ValueError(f"Errores al validar eventos Excel:\n{mensajes}")

    return eventos


def _fila_vacia(ws, row: int) -> bool:
    """Devuelve True si todas las celdas de la fila están vacías."""
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            return False
    return True
